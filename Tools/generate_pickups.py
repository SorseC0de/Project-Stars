#!/usr/bin/env python3
"""Turns _Design/Pickups.xlsx into Swift.

The sheet is the source of truth for every pickup's *descriptor* — its name,
its rules line, how often it turns up. What a pickup *does* stays in Swift,
because it is behaviour and a spreadsheet cannot hold it.

Run from the `Project Stars` folder:

    python3 Tools/generate_pickups.py

It writes ProjectStars/Pickups/PickupDescriptors.generated.swift and refuses to
write anything if the sheet disagrees with the Swift it is generating against —
a row whose struct does not exist, or a struct with no row. Silence on either
would mean a coin quietly losing its name.
"""

import pathlib
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is not installed.  pip3 install openpyxl")

ROOT = pathlib.Path(__file__).resolve().parent.parent
SHEET = ROOT / "_Design" / "Pickups.xlsx"
SOURCE = ROOT / "ProjectStars" / "Pickups" / "Pentacles.swift"
OUT = ROOT / "ProjectStars" / "Pickups" / "PickupDescriptors.generated.swift"

COLUMNS = ["struct", "id", "displayName", "summary", "rarity", "rollsAsRarity",
           "chance", "glyph", "icon", "spawnPlane", "element", "appearance",
           "pickupClass", "notes"]


def literal(text):
    """A Swift string literal that keeps its interpolation.

    An empty cell arrives as `None`, and `str(None)` is the word "None" — which
    is how two coins briefly came to be called that.

    Only the quote is escaped. A backslash in one of these cells is there to
    interpolate — `Gain \\(GameRules.zChargePentacleAmount) ZC` — and escaping it
    would print the machinery instead of the number.
    """
    return '"' + str(text or "").replace('"', '\\"') + '"'


def case(value, kind):
    """A Swift enum case, or `nil` for a blank cell."""
    value = (value or "").strip()
    return f".{value}" if value else ("nil" if kind == "optional" else None)


def wrap(text, width=72):
    """The notes column, as comment lines."""
    words = str(text or "").split()
    if not words:
        return []
    lines, line = [], words[0]
    for word in words[1:]:
        if len(line) + 1 + len(word) > width:
            lines.append(line)
            line = word
        else:
            line += " " + word
    lines.append(line)
    return lines


def read_rows():
    book = load_workbook(SHEET, data_only=True)
    sheet = book["Pickups"]

    header = [c.value for c in sheet[1]]
    if header[: len(COLUMNS)] != ["Struct", "ID", "Name", "Summary", "Rarity",
                                  "Rolls As", "Chance", "Glyph", "Icon", "Plane",
                                  "Element", "Appearance", "Class", "Notes"]:
        sys.exit(f"{SHEET.name}: unexpected columns {header}")

    rows = []
    for line in sheet.iter_rows(min_row=2, values_only=True):
        if not line or not line[0]:
            break                      # the blank line before the total
        rows.append(dict(zip(COLUMNS, line)))
    return rows


def main():
    if not SHEET.exists():
        sys.exit(f"no sheet at {SHEET}")

    rows = read_rows()
    swift = SOURCE.read_text()
    declared = set(re.findall(r"struct (\w+Effect): PickupEffect \{", swift))
    listed = {r["struct"] for r in rows}

    if orphans := sorted(listed - declared):
        sys.exit("rows with no struct in Pentacles.swift: " + ", ".join(orphans))
    if missing := sorted(declared - listed):
        sys.exit("structs with no row in the sheet: " + ", ".join(missing))

    lines = [
        "//",
        "//  PickupDescriptors.generated.swift",
        "//  Project Stars",
        "//",
        "//  GENERATED FROM _Design/Pickups.xlsx — DO NOT EDIT.",
        "//  Change the sheet and run: python3 Tools/generate_pickups.py",
        "//",
        "",
        "/// What every pickup is called, says, and how often it turns up.",
        "///",
        "/// One table rather than a property on each of twenty-seven structs,",
        "/// because it is one kind of fact and it is authored somewhere else. The",
        "/// structs keep `plan(...)`, which is the half a spreadsheet cannot hold.",
        "enum PickupDescriptors {",
        "",
        "    static let all: [PickupID: PickupDescriptor] = [",
    ]

    for row in sorted(rows, key=lambda r: r["struct"]):
        # The reasoning, back above the number it is about. It was written in
        # Swift, it lives in the sheet now, and it is put back here so that
        # whichever of the two you are reading, the number is explained.
        for note in wrap(row["notes"]):
            lines.append(f"        // {note}")

        chance = row["chance"]
        chance = str(chance) if isinstance(chance, int) else str(chance).strip()
        icon = row["icon"]
        lines += [
            f'        .{row["id"]}: PickupDescriptor(',
            f'            displayName: {literal(row["displayName"])},',
            f'            summary: {literal(row["summary"])},',
            f'            glyph: {literal(row["glyph"])},',
            f'            icon: {literal(icon) if icon else "nil"},',
            f'            rarity: {case(row["rarity"], "required")},',
            f'            rollsAsRarity: {case(row["rollsAsRarity"], "optional")},',
            f'            chance: {chance},',
            f'            spawnPlane: {case(row["spawnPlane"], "optional")},',
            f'            element: {case(row["element"], "optional")},',
            f'            appearance: {case(row["appearance"], "required") or ".standard"},',
            f'            pickupClass: {case(row["pickupClass"], "required") or ".pentacle"}',
            "        ),",
        ]

    lines += ["    ]", "}", ""]
    OUT.write_text("\n".join(lines))
    print(f"wrote {OUT.relative_to(ROOT)} — {len(rows)} pickups")


if __name__ == "__main__":
    main()
